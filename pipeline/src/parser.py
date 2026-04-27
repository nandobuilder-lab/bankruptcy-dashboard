"""
Parsers for US Courts bankruptcy filing XLSX tables.

Table types handled:
  bf_f_MMDD.YYYY.xlsx    → Table F  (Filed/Terminated/Pending, 12-month, by district)
  bf_f2_MMDD.YYYY.xlsx   → Table F-2 (Filings by chapter, 12-month, by district)
  bf_f2.1_MMDD.YYYY.xlsx → Table F-2.1 (same, 1-month)
  bf_f2.3_MMDD.YYYY.xlsx → Table F-2.3 (same, 3-month)
  bf_f5a_MMDD.YYYY.xlsx  → Table F-5A (Filings by chapter, district, and county)

All parsers return a pandas DataFrame ready for loading into DuckDB.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Cell value normalisation
# ---------------------------------------------------------------------------

_FOOTNOTE_RE = re.compile(r"[¹²³⁴⁵⁶⁷⁸⁹\*]")


def _int(val) -> Optional[int]:
    """Coerce a cell value to int, returning None for suppressed/missing cells."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return None if pd.isna(val) else int(val)
    s = _FOOTNOTE_RE.sub("", str(val)).strip().replace(",", "")
    if s in ("", "-", "N/A", "n/a"):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _row_type(raw_label: str) -> str:
    """Classify a district-column value as national / circuit / district."""
    stripped = raw_label.strip()
    if stripped.upper() == "TOTAL":
        return "national"
    if raw_label != raw_label.lstrip():   # leading whitespace → circuit header
        return "circuit"
    return "district"


# ---------------------------------------------------------------------------
# Filename metadata
# ---------------------------------------------------------------------------

_FILENAME_RE = re.compile(r"bf_(f[\w.]*?)_(\d{2})(\d{2})\.(\d{4})\.xlsx$")


def _meta_from_filename(filename: str) -> Tuple[str, date]:
    """
    Parse table type and period-end date from a filename.

    Returns:
        table_type: one of 'f', 'f2', 'f2.1', 'f2.3', 'f5a'
        period_end: date object
    """
    m = _FILENAME_RE.search(filename)
    if not m:
        raise ValueError(f"Cannot parse table metadata from filename: {filename!r}")
    table_type = m.group(1)
    mm, dd, yyyy = int(m.group(2)), int(m.group(3)), int(m.group(4))
    return table_type, date(yyyy, mm, dd)


def period_months_from_table(table_type: str) -> int:
    return {"f2.1": 1, "f2.3": 3}.get(table_type, 12)


# ---------------------------------------------------------------------------
# Table F parser
# ---------------------------------------------------------------------------

def parse_f(path: Path, release_id: str) -> pd.DataFrame:
    """
    Parse Table F (Filed / Terminated / Pending) into a flat DataFrame.

    Columns:
        release_id, period_end, row_type, label,
        prior_year, current_year,
        filed_prior, filed_current,
        terminated_prior, terminated_current,
        pending_prior, pending_current
    """
    _, period_end = _meta_from_filename(path.name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Year labels row position varies by format (old: row 3, new: row 4, 0-indexed).
    # Scan rows 2–5 for the first row where col A is empty and col B looks like a year.
    year_row_idx = 3
    for i, r in enumerate(rows[2:6], start=2):
        y = _int(r[1]) if r[1] is not None else None
        if r[0] is None and y is not None and 2000 <= y <= 2040:
            year_row_idx = i
            break
    year_row = rows[year_row_idx]
    prior_year = _int(year_row[1])
    current_year = _int(year_row[2])

    records = []
    for row in rows[year_row_idx + 1:]:
        if row[0] is None:
            continue
        raw = str(row[0])
        label = raw.strip()
        if not label or label.upper().startswith("NOTE") or label[0] in "¹²³⁴⁵⁶⁷⁸⁹":
            break

        records.append({
            "release_id":          release_id,
            "period_end":          period_end,
            "row_type":            _row_type(raw),
            "label":               label,
            "prior_year":          prior_year,
            "current_year":        current_year,
            "filed_prior":         _int(row[1]),
            "filed_current":       _int(row[2]),
            "terminated_prior":    _int(row[4]),
            "terminated_current":  _int(row[5]),
            "pending_prior":       _int(row[7]),
            "pending_current":     _int(row[8]),
        })

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Table F-2 / F-2.1 / F-2.3 parsers
# ---------------------------------------------------------------------------

_PERIOD_END_RE = re.compile(r"Ending\s+(\w+ \d+,\s*\d{4})", re.IGNORECASE)


def _period_end_from_title(title: str) -> Optional[date]:
    """Extract the period-end date from the XLSX title cell text."""
    m = _PERIOD_END_RE.search(title)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1).replace(",  ", ", "), "%B %d, %Y").date()
    except ValueError:
        return None


def _parse_f2_sheet(ws, release_id: str, period_end: date, period_months: int) -> List[dict]:
    """Parse one F-2 worksheet (15 columns) into a list of record dicts."""
    records = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[4:]:    # data at row 5 (old format) or row 6 (new); None rows skipped
        if row[0] is None:
            continue
        raw = str(row[0])
        label = raw.strip()
        if not label or label.upper().startswith("NOTE") or label[0] in "¹²³⁴⁵⁶⁷⁸⁹":
            break
        records.append({
            "release_id":    release_id,
            "period_end":    period_end,
            "period_months": period_months,
            "row_type":      _row_type(raw),
            "label":         label,
            "total_all":     _int(row[1]),
            "total_ch7":     _int(row[2]),
            "total_ch11":    _int(row[3]),
            "total_ch13":    _int(row[4]),
            "total_other":   _int(row[5]),
            "biz_all":       _int(row[6]),
            "biz_ch7":       _int(row[7]),
            "biz_ch11":      _int(row[8]),
            "biz_ch13":      _int(row[9]),
            "biz_other":     _int(row[10]),
            "nonbiz_all":    _int(row[11]),
            "nonbiz_ch7":    _int(row[12]),
            "nonbiz_ch11":   _int(row[13]),
            "nonbiz_ch13":   _int(row[14]) if len(row) > 14 else None,
        })
    return records


def parse_f2(path: Path, release_id: str) -> pd.DataFrame:
    """
    Parse Table F-2 (12-month) or F-2.3 (3-month) — single active sheet.

    period_months is inferred from the filename: f2_ → 12, f2.3_ → 3.

    Columns:
        release_id, period_end, period_months, row_type, label,
        total_{all,ch7,ch11,ch13,other},
        biz_{all,ch7,ch11,ch13,other},
        nonbiz_{all,ch7,ch11,ch13}
    """
    table_type, period_end = _meta_from_filename(path.name)
    period_months = period_months_from_table(table_type)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = _parse_f2_sheet(ws, release_id, period_end, period_months)
    return pd.DataFrame(records)


def parse_f2_monthly(path: Path, release_id: str) -> pd.DataFrame:
    """
    Parse Table F-2.1 (1-month) — one sheet per month within the quarter.

    Each quarterly file contains 6 sheets: 3 main sheets (15 cols, standard
    business/nonbusiness breakdown) and 3 rare-chapters sheets (6 cols,
    chapters 9/12/15 only). Only the main sheets are parsed here.

    The period-end date for each month is extracted from the sheet's title cell.

    Returns a single DataFrame with period_months=1, one row per district per month.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    all_records: List[dict] = []

    for sname in wb.sheetnames:
        if "(9, 12, 15)" in sname:
            continue                    # skip rare-chapters sheets
        ws = wb[sname]
        if ws.max_column < 10:          # safety: skip any unexpectedly narrow sheets
            continue

        # Extract period_end from the title cell (row 2 in new format, row 1 in old format)
        rows = list(ws.iter_rows(values_only=True))
        title = str(rows[1][0]) if rows[1][0] else ""
        period_end = _period_end_from_title(title)
        if period_end is None:
            title = str(rows[0][0]) if rows[0][0] else ""
            period_end = _period_end_from_title(title)
        if period_end is None:
            continue                    # can't determine date → skip

        all_records.extend(_parse_f2_sheet(ws, release_id, period_end, 1))

    return pd.DataFrame(all_records)


# ---------------------------------------------------------------------------
# Table F-5A parser
# ---------------------------------------------------------------------------

def parse_f5a(path: Path, release_id: str) -> pd.DataFrame:
    """
    Parse Table F-5A (county-level) into a flat DataFrame.

    Rows are: national total → circuit totals → district totals → county rows.
    The parser is stateful: it tracks current circuit/district to annotate counties.

    FIPS codes ending with '*' indicate the filing was out-of-district
    (filed in this district but geographically in another).

    Columns:
        release_id, period_end, row_type, circuit, district, label,
        county_name, county_fips, is_out_of_district,
        total_{all,ch7,ch11,ch13,other},
        biz_{all,ch7,ch11,ch13,other},
        nonbiz_{all,ch7,ch11,ch13}
    """
    _, period_end = _meta_from_filename(path.name)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    current_circuit: Optional[str] = None
    current_district: Optional[str] = None
    records = []

    for row in rows[4:]:    # data at row 5 (old format) or row 6 (new); None rows skipped
        if row[0] is None:
            continue
        raw = str(row[0])
        label = raw.strip()
        if not label:
            continue                 # old-format files use whitespace-only filler rows
        if label.upper().startswith("NOTE") or label[0] in "¹²³⁴⁵⁶⁷⁸⁹":
            break

        county_raw = str(row[1]).strip() if row[1] is not None else None
        if county_raw in (None, "None", ""):
            county_raw = None

        if label == "Total":
            rt = "national"
            current_circuit = current_district = None
            county_name = county_fips = None
            is_ood = False
        elif raw != label:          # leading whitespace → circuit
            rt = "circuit"
            current_circuit = label
            current_district = None
            county_name = county_fips = None
            is_ood = False
        elif county_raw is None:    # no FIPS → district aggregate
            rt = "district"
            current_district = label
            county_name = county_fips = None
            is_ood = False
        else:                       # county row
            rt = "county"
            county_name = label
            is_ood = county_raw.endswith("*")
            county_fips = county_raw.rstrip("*")

        def c(i: int) -> int | None:
            return _int(row[i]) if len(row) > i else None

        records.append({
            "release_id":         release_id,
            "period_end":         period_end,
            "row_type":           rt,
            "circuit":            current_circuit,
            "district":           current_district,
            "label":              label,
            "county_name":        county_name,
            "county_fips":        county_fips,
            "is_out_of_district": is_ood,
            "total_all":          c(2),
            "total_ch7":          c(3),
            "total_ch11":         c(4),
            "total_ch13":         c(5),
            "total_other":        c(6),
            "biz_all":            c(7),
            "biz_ch7":            c(8),
            "biz_ch11":           c(9),
            "biz_ch13":           c(10),
            "biz_other":          c(11),
            "nonbiz_all":         c(12),
            "nonbiz_ch7":         c(13),
            "nonbiz_ch11":        c(14),
            "nonbiz_ch13":        c(15) if len(row) > 15 else None,
        })

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------

def parse_file(path: Path, release_id: str) -> Tuple[str, pd.DataFrame]:
    """
    Parse a downloaded XLSX file and return (db_table_name, dataframe).

    db_table_name is one of: 'f_summary', 'f2_filings', 'f5a_county'
    """
    name = path.name
    if name.startswith("bf_f5a"):
        return "f5a_county", parse_f5a(path, release_id)
    if name.startswith("bf_f2.1"):
        return "f2_filings", parse_f2_monthly(path, release_id)
    if name.startswith("bf_f2"):
        return "f2_filings", parse_f2(path, release_id)
    if name.startswith("bf_f_"):
        return "f_summary", parse_f(path, release_id)
    raise ValueError(f"Unknown table file: {name!r}")
